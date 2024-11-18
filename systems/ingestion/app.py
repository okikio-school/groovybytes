from flask import Flask, request, jsonify, render_template, render_template_string, send_file
from werkzeug.utils import secure_filename
from apscheduler.schedulers.background import BackgroundScheduler
from paho.mqtt.client import Client as MQTTClient
import pandas as pd
import os
import json
import requests
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
import datetime

UPLOAD_FOLDER = 'uploads'
OUTPUT_FILE = 'output_data.xlsx'
HISTORY_FILE = 'upload_history.json'
OUTPUT_URL = 'http://localhost:5001/formatting/process'
ALLOWED_EXTENSIONS = {'csv', 'xls', 'xlsx', 'json'}
file_queue = []
file_status = {}

app = Flask(__name__)
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER

mqtt_client = MQTTClient()
MQTT_BROKER = "broker.emqx.io"  # broker.hivemq.com   broker.emqx.io
MQTT_TOPIC = "sensor/temperature"

scheduler = BackgroundScheduler()
scheduler.start()


def load_history():
    if os.path.exists(HISTORY_FILE):
        with open(HISTORY_FILE, 'r') as f:
            return json.load(f)
    return {}


def save_history(history):
    with open(HISTORY_FILE, 'w') as f:
        json.dump(history, f, indent=4)


file_status = load_history()


def on_connect(client, userdata, flags, rc):
    print("Connected to MQTT Broker")

    client.subscribe(MQTT_TOPIC)


def on_message(client, userdata, msg):
    print(f"Received message from MQTT: {msg.payload.decode()}")
    data = msg.payload.decode()



def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS


@app.route('/')
def index():
    return render_template('index.html')


@app.route('/ingestion/upload', methods=['POST'])
def upload_file():
    if 'file' not in request.files:
        return jsonify({'status': 'fail', 'message': 'No file uploaded'})

    file = request.files['file']
    if file.filename == '':
        return jsonify({'status': 'fail', 'message': 'No file selected'})

    if file and allowed_file(file.filename):
        filename = secure_filename(file.filename)
        file_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)
        file.save(file_path)

        file_status[filename] = 'uploaded'
        file_queue.append(file_path)

        save_history(file_status)

        return jsonify({'status': 'success', 'message': 'File uploaded successfully and queued for processing',
                        'file_path': file_path}, 200)
    else:
        return jsonify({'status': 'fail', 'message': 'Invalid file type'})


def append_to_excel(data, source_label):
    # Check if the output file exists, and load it; otherwise, create a new workbook
    if os.path.exists(OUTPUT_FILE):
        workbook = load_workbook(OUTPUT_FILE)
    else:
        workbook = Workbook()
        workbook.save(OUTPUT_FILE)
        workbook = load_workbook(OUTPUT_FILE)

    sheet = workbook.active
    next_row = sheet.max_row + 2 if sheet.max_row > 1 else 1  # Leave a blank row between entries

    # Add source label (filename or MQTT source) in bold in the first column
    sheet.cell(row=next_row, column=1, value=source_label).font = Font(bold=True)

    # Set the first row of CSV as headers in Excel if not set
    headers = data.columns
    for col_num, col_name in enumerate(headers, start=2):  # Start from second column for headers
        sheet.cell(row=next_row + 1, column=col_num, value=col_name).font = Font(bold=True)

    # Move the row pointer down after setting headers
    next_row += 1

    # Append data with formatted content: cell_content::source_label::COLUMN_NAME
    for row in data.itertuples(index=False, name=None):
        for col_num, (value, col_name) in enumerate(zip(row, data.columns), start=2):  # Start from second column
            formatted_value = f"{value}::{source_label}::{col_name}"
            sheet.cell(row=next_row + 1, column=col_num, value=formatted_value)
        next_row += 1

    # Save the modified workbook
    workbook.save(OUTPUT_FILE)


def process_file(file_path):
    filename = os.path.basename(file_path)
    file_status[filename] = 'processing'
    save_history(file_status)

    try:
        _, ext = os.path.splitext(file_path)

        if ext == '.csv':
            data = pd.read_csv(file_path)
        elif ext in ['.xls', '.xlsx']:
            data = pd.read_excel(file_path)
        elif ext == '.json':
            data = pd.read_json(file_path)
        else:
            print("Unsupported file type:", file_path)
            file_status[filename] = 'error'
            save_history(file_status)
            return

        send_to_output_sink(data)

        file_status[filename] = 'processed'
        save_history(file_status)

    except Exception as e:
        print(f"Error processing file {file_path}: {e}")
        file_status[filename] = 'error'
        save_history(file_status)


def process_mqtt_data(data):
    try:
        df = pd.DataFrame([json.loads(data)])
        timestamp_label = f"MQTT {datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"

        file_status[timestamp_label] = 'processing'
        save_history(file_status)

        append_to_excel(df, timestamp_label)
        file_status[timestamp_label] = 'processed'
        save_history(file_status)

        print(f"MQTT data processed and appended to {OUTPUT_FILE}")

    except Exception as e:
        print(f"Error processing MQTT data: {e}")
        file_status[timestamp_label] = 'error'
        save_history(file_status)


def process_files_in_queue():
    while file_queue:
        file_to_process = file_queue.pop(0)
        process_file(file_to_process)

def send_to_output_sink(data):
    # CODE TO SEND TO OTHER SERVER
    # URL of the endpoint
    # Make the POST request
    try:
        json_data = data.to_json(orient='records')
        response = requests.post(OUTPUT_URL, json=json_data)

        # Check the response
        if response.status_code == 200:
            print("Response from server:", response.json())
        else:
            print(f"Failed to send data. Status code: {response.status_code}, Message: {response.text}")
    except Exception as e:
        print(f"Error making POST request: {e}")


scheduler.add_job(process_files_in_queue, 'interval', seconds=10, max_instances=2)


@app.route('/view_data')
def view_data():
    if not os.path.exists(OUTPUT_FILE):
        return "<h3>No data available. Upload files to view data.</h3>"

    df = pd.read_excel(OUTPUT_FILE)

    # Replace "Unnamed" headers with empty strings
    df.columns = ["" if "Unnamed" in col else col for col in df.columns]

    # Replace NaN values with empty strings for a clean display in HTML
    df.fillna('', inplace=True)

    # Convert DataFrame to HTML table with Bootstrap classes
    html_table = df.to_html(index=False, classes="table table-striped table-bordered table-hover", na_rep="")

    return render_template_string('''
    <link rel="stylesheet" href="https://stackpath.bootstrapcdn.com/bootstrap/4.5.2/css/bootstrap.min.css">
    <style>
        table { font-size: small; }
        th { background-color: #f8f9fa; font-weight: bold; }
        td, th { text-align: left;  vertical-align: middle; padding: 0.5rem;}
        td:first-child { font-weight: bold; }
    </style>
    <div class="container-fluid">
        <div class="row align-items-center my-4">
            <div class="col-md-4">
                <a href="/" class="btn btn-secondary">Go Back</a>
            </div>
            <div class="col-md-4 text-center">
                <h1>Output Sink</h1>
            </div>
            <div class="col-md-4 text-right">
                <a href="{{ url_for('static', filename='output_data.xlsx') }}" class="btn btn-success">
                    Download
                </a>
            </div>
        </div>
        <div class="table-responsive">
            {{ table | safe }}
        </div>
    </div>
    ''', table=html_table)


@app.route('/download_output')
def download_output():
    if os.path.exists(OUTPUT_FILE):
        return send_file(OUTPUT_FILE, as_attachment=True)
    else:
        return "<h3>No output file available to download.</h3>"


@app.route('/status')
def status():
    return jsonify(file_status)


if __name__ == '__main__':
    os.makedirs(UPLOAD_FOLDER, exist_ok=True)

    mqtt_client.on_connect = on_connect
    mqtt_client.on_message = on_message
    mqtt_client.connect(MQTT_BROKER, 1883, 60)
    mqtt_client.loop_start()

    app.run(port=5000)